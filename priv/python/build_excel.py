#!/usr/bin/env python3
"""
Script to convert the stick diagram PDF into an Excel file with the same data.
"""
import openpyxl
import json
import os
import sys

from copy import copy
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(SCRIPT_DIR, "excel_templates")

templates = [
    os.path.join(TEMPLATE_DIR, 'TEMPLATE_EOW Report.xlsx'),
    os.path.join(TEMPLATE_DIR, 'TEMPLATE_AM Report.xlsx'),
    os.path.join(TEMPLATE_DIR, 'TEMPLATE2_AM Report.xlsx'),
    os.path.join(TEMPLATE_DIR, 'TEMPLATE_Mud Resistivity.xlsx'),
    os.path.join(TEMPLATE_DIR, 'TEMPLATE_Sample Descriptions.xlsx'),
]

def populate_eow_report(data, output_dir):
    wb = openpyxl.load_workbook(os.path.join(TEMPLATE_DIR, 'TEMPLATE_EOW Report.xlsx'))
    ws = wb.active

    ws['A3'] = data['well_name']
    ws['D3'] = data['uwi']
    ws['H3'] = data['licence']
    ws['H5'] = data.get('rig_name', '')
    ws['I3'] = data.get('spud_date', '')
    ws['F3'] = data.get('geo_day', '') + ' (Day) \n ' + data.get('geo_night', '') + ' (Night)'

    out_path = os.path.join(output_dir, data["well_name"] + "_EOW Report.xlsx")
    wb.save(out_path)
    return out_path

def populate_am_report(data, output_dir):
    wb = openpyxl.load_workbook(os.path.join(TEMPLATE_DIR, 'TEMPLATE2_AM Report.xlsx'))
    ws = wb.active

    mcmurray_top = 0
    total_depth = 0
    last_entry_row = 0
    formations_info_length = len(data['sections']['Geological Formation Information']['rows'])

    # Write geological formation rows starting at row 34
    row = 34
    for formation in data['sections']['Geological Formation Information']['rows']:
        ws[f'A{row}'] = formation.get('formation')
        ws[f'D{row}'] = formation.get('mkb_tvd')
        ws[f'E{row}'] = formation.get('mkb_tvd')
        ws[f'F{row}'] = formation.get('masl_tvd')
        name = (formation.get('formation') or '').lower()

        if 'mcmurray' in name and mcmurray_top == 0:
            mcmurray_top = formation.get('mkb_tvd') or 0
        if formation.get('mkb_tvd') is not None:
            total_depth = formation.get('mkb_tvd')
        
        last_entry_row = row
        row += 1

    # Build general information
    ws['C8'] = data.get('rig_name', '')
    ws['C11'] = data['licence']
    ws['D2'] = data['well_name']
    ws['D3'] = data['uwi']
    ws['C12'] = data.get('afe', '')
    ws['C13'] = data.get('spud_date', '')
    ws['I5'] = data.get('og', '')
    ws['L5'] = data.get('og_ph', '')
    ws['I6'] = data.get('geo_day', '')
    ws['L6'] = data.get('geo_day_ph', '')
    ws['I7'] = data.get('geo_night', '')
    ws['L7'] = data.get('geo_night_ph', '')
    ws['I8'] = data.get('wss_day', '')
    ws['I9'] = data.get('wss_night', '')
    ws['L9'] = data.get('rig_ph', '')

    # Build geological formation information
    if last_entry_row < 52:
        rows_to_delete = 52 - last_entry_row

        # Collect and shift merged cell ranges before deleting
        new_merges = []
        delete_start = last_entry_row + 1
        delete_end = last_entry_row + rows_to_delete

        for merge in ws.merged_cells.ranges:
            # Skip any merge that overlaps the rows being deleted
            if merge.max_row >= delete_start and merge.min_row <= delete_end:
                continue
            
            if merge.min_row > delete_end:
                # Below the deletion zone — shift up
                new_merges.append((
                    merge.min_row - rows_to_delete,
                    merge.min_col,
                    merge.max_row - rows_to_delete,
                    merge.max_col
                ))
            else:
                new_merges.append((
                    merge.min_row,
                    merge.min_col,
                    merge.max_row,
                    merge.max_col
                ))

        ws.merged_cells.ranges.clear()
        ws.delete_rows(last_entry_row + 1, rows_to_delete)

        for min_row, min_col, max_row, max_col in new_merges:
            ws.merge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col
            )

    # Well Geometry section L12:N12


    # Resistivity section
    smpl_depth = 34
    start = mcmurray_top - 20
    while start <= total_depth:
        ws[f'L{smpl_depth}'] = start
        start += 50
        smpl_depth += 1
    ws[f'L{smpl_depth}'] = 'TD'
    
    out_path = os.path.join(output_dir, data["well_name"] + "_AM Report.xlsx")
    wb.save(out_path)
    return out_path, mcmurray_top, total_depth


def populate_resistivity_report(data, output_dir, mcmurray_top, total_depth):
    wb = openpyxl.load_workbook(os.path.join(TEMPLATE_DIR, 'TEMPLATE_Mud Resistivity.xlsx'))
    ws = wb.active
    row = 6
    start = mcmurray_top - 20
    
    ws['A3'] = data['well_name']
    ws['A4'] = data['uwi']

    while start <= total_depth:
        ws[f'B{row}'] = start
        start += 50
        row += 1
    ws[f'B{row}'] = 'TD'

    out_path = os.path.join(output_dir, data["well_name"] + "_Mud Resistivity.xlsx")
    wb.save(out_path)
    return out_path


def main():
    mcmurray_top = 0
    total_depth = 0
    data = json.loads(sys.argv[1] if len(sys.argv) > 1 else sys.stdin.read())
    output_dir = data.get("output_dir", os.getcwd())

    files = []
    files.append(populate_eow_report(data, output_dir))
    am_report, mcmurray_top, total_depth = populate_am_report(data, output_dir)
    files.append(am_report)
    resistivity_report = populate_resistivity_report(data, output_dir, mcmurray_top, total_depth)
    files.append(resistivity_report)

    print(json.dumps({"status": "ok", "files": files}))


if __name__ == "__main__":
    main()